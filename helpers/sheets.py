"""
Google Sheets ingestion for the raid-setups feature.

Reads a **publicly shared** Google Sheet (share = "Anyone with the link →
Viewer") through Google's plain XLSX export endpoint, so it works for anyone's
sheet without OAuth, service accounts, or a personal Drive connector — the only
requirement is link-sharing. The workbook is parsed into a plain data model the
``raid_setups`` cog stores in MongoDB.

Expected template (see the Instructions tab of the shared template):
  * ``Roster`` tab  — headers ``Role | Name | Class | Slayer | Notes`` (+ any
    extra columns), one row per player.
  * ``Setups`` tab  — stacked stage blocks. Each block is a stage-name row (only
    the first column filled), a ``Player | Set 1 | Set 2 | …`` header row, then
    one row per player, blocks separated by a blank row. Any number of blocks,
    named anything, in any order.

Kept free of Discord imports so it can be unit-tested on its own.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

import openpyxl
import requests

ROSTER_TAB = "Roster"
SETUPS_TAB = "Setups"
# First-column header labels that mark a stage block's header row (lower-cased).
_HEADER_LABELS = {"player", "role", "name"}
# Header labels for the "bold this setup" checkbox column (lower-cased).
_BOLD_LABELS = {"★", "✓", "✔", "bold", "highlight", "current", "active", "show"}
# Reserved key used to carry a row's bold flag inside its values dict.
_BOLD_KEY = "__bold__"


def _truthy(value) -> bool:
    """Interpret a checkbox cell (bool, TRUE/FALSE text, 1/0, x, ✓) as a boolean."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"true", "1", "yes", "y", "x", "✓", "✔", "checked"}

# Roster columns the bot understands (matched case-insensitively by header name).
# "Discord" holds the player's Discord username/tag and gates the /setups command.
_ROSTER_FIELDS = ("Role", "Name", "Discord", "Class", "Slayer", "Notes")

_SHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9_-]+)")
_BARE_ID_RE = re.compile(r"^[a-zA-Z0-9_-]{20,}$")


class SheetError(Exception):
    """Raised for a bad link, an unreachable/unshared sheet, or a malformed template."""


@dataclass
class Stage:
    """One pull/stage block: an ordered set of per-player gear rows."""

    name: str
    order: int
    columns: list[str]                      # gear columns after "Player", e.g. Set 1, Set 2, …
    rows: dict[str, dict[str, str]]         # player name -> {column: value}


@dataclass
class RaidData:
    """The parsed contents of a raid sheet."""

    roster: list[dict[str, str]] = field(default_factory=list)
    stages: list[Stage] = field(default_factory=list)
    columns: list[str] = field(default_factory=list)   # union of stage columns, first-seen order
    warnings: list[str] = field(default_factory=list)

    @property
    def player_names(self) -> list[str]:
        """Roster player names, in roster order."""
        return [row["Name"] for row in self.roster if row.get("Name")]

    def roster_entry(self, name: str) -> dict[str, str] | None:
        """Case-insensitive roster lookup by player name."""
        for row in self.roster:
            if row.get("Name", "").lower() == name.lower():
                return row
        return None

    def match_discord(self, candidates: set[str]) -> dict[str, str] | None:
        """Find the roster row whose Discord tag matches one of ``candidates``.

        ``candidates`` are the caller's normalised identities (user id as string,
        username, global name), all lower-cased. The sheet's Discord cell is
        compared after stripping a leading ``@``.
        """
        for row in self.roster:
            tag = row.get("Discord", "").strip().lstrip("@").lower()
            if tag and tag in candidates:
                return row
        return None

    def lookup(self, name: str) -> list[tuple[str, dict[str, str], bool]]:
        """Return ``[(stage_name, {column: value}, bold), …]`` for one player.

        ``bold`` is the stage's checkbox flag for that player. Stages where the
        player has no row are included with empty values, so the pull list stays
        complete.
        """
        result = []
        for stage in self.stages:
            values = stage.rows.get(name) or _ci_get(stage.rows, name) or {}
            row = {col: values.get(col, "") for col in self.columns}
            result.append((stage.name, row, bool(values.get(_BOLD_KEY))))
        return result


    def to_mongo(self) -> dict:
        """Serialise the parsed data for storage (roster + stages + columns)."""
        return {
            "columns": self.columns,
            "roster": self.roster,
            "stages": [
                {"name": s.name, "order": s.order, "columns": s.columns, "rows": s.rows}
                for s in self.stages
            ],
        }

    @classmethod
    def from_mongo(cls, doc: dict) -> "RaidData":
        """Rebuild ``RaidData`` from a stored raid document."""
        data = cls()
        data.columns = doc.get("columns", [])
        data.roster = doc.get("roster", [])
        data.stages = [
            Stage(name=s["name"], order=s.get("order", i), columns=s.get("columns", []), rows=s.get("rows", {}))
            for i, s in enumerate(doc.get("stages", []))
        ]
        return data


def _ci_get(mapping: dict[str, dict[str, str]], key: str):
    """Case-insensitive dict lookup helper."""
    for k, v in mapping.items():
        if k.lower() == key.lower():
            return v
    return None


def extract_sheet_id(url_or_id: str) -> str:
    """Pull the sheet ID out of any Google Sheets URL, or accept a bare ID."""
    text = (url_or_id or "").strip()
    match = _SHEET_ID_RE.search(text)
    if match:
        return match.group(1)
    if _BARE_ID_RE.match(text):
        return text
    raise SheetError(
        "That doesn't look like a Google Sheets link. Paste the full URL from your "
        "browser (it contains `/spreadsheets/d/…`)."
    )


def fetch_workbook(sheet_id: str, *, timeout: float = 15.0) -> openpyxl.Workbook:
    """Download a public sheet as XLSX and open it. Raises ``SheetError`` if not reachable."""
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        response = requests.get(export_url, timeout=timeout)
    except requests.RequestException as error:
        raise SheetError(f"Couldn't reach Google Sheets: {error}") from error

    ctype = response.headers.get("content-type", "")
    if response.status_code != 200 or "spreadsheet" not in ctype:
        raise SheetError(
            "Couldn't read that sheet. Make sure sharing is set to "
            "**Anyone with the link → Viewer**, then try again."
        )
    try:
        return openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
    except Exception as error:  # noqa: BLE001 - openpyxl raises assorted errors
        raise SheetError(f"That file isn't a readable spreadsheet: {error}") from error


def parse_workbook(workbook: openpyxl.Workbook) -> RaidData:
    """Parse a loaded workbook into ``RaidData``. Raises ``SheetError`` on fatal issues."""
    tabs = {name.lower(): name for name in workbook.sheetnames}
    if ROSTER_TAB.lower() not in tabs:
        raise SheetError(f"The sheet is missing a **{ROSTER_TAB}** tab.")
    if SETUPS_TAB.lower() not in tabs:
        raise SheetError(f"The sheet is missing a **{SETUPS_TAB}** tab.")

    data = RaidData()
    data.roster = _parse_roster(workbook[tabs[ROSTER_TAB.lower()]])
    data.stages, data.columns, stage_warnings = _parse_setups(workbook[tabs[SETUPS_TAB.lower()]], data.roster)
    data.warnings.extend(stage_warnings)

    if not data.roster:
        raise SheetError(f"The **{ROSTER_TAB}** tab has no players.")
    if not data.stages:
        raise SheetError(f"The **{SETUPS_TAB}** tab has no stage blocks.")

    data.warnings.extend(_cross_check(data))
    return data


def _cells(row) -> list[str]:
    """Normalise a row of cell values to trimmed strings."""
    return ["" if value is None else str(value).strip() for value in row]


def _parse_roster(ws) -> list[dict[str, str]]:
    """Parse the Roster tab into a list of player dicts keyed by canonical field names."""
    rows = ws.iter_rows(values_only=True)
    header = None
    for raw in rows:
        cells = _cells(raw)
        if any(cells):
            header = cells
            break
    if not header:
        return []

    # Map each known field to its column index by case-insensitive header match.
    index = {}
    for field_name in _ROSTER_FIELDS:
        for col, title in enumerate(header):
            if title.lower() == field_name.lower():
                index[field_name] = col
                break

    roster = []
    for raw in rows:
        cells = _cells(raw)
        if not any(cells):
            continue
        name_col = index.get("Name")
        if name_col is None or name_col >= len(cells) or not cells[name_col]:
            continue  # rows without a name aren't players
        entry = {}
        for field_name, col in index.items():
            entry[field_name] = cells[col] if col < len(cells) else ""
        roster.append(entry)
    return roster


def _parse_setups(ws, roster: list[dict[str, str]]):
    """Parse the Setups tab's stacked stage blocks.

    A block is: a stage-title row (a single cell with the pull's name), a header
    row (``Player | Set 1 | …``), and one row per player, blocks separated by a
    blank row. Only **titled** blocks are imported, so the template can ship with
    many pre-made, untitled blocks that stay dormant until named.

    Player identity is the roster **Name**: each row's first cell is matched to
    the roster (case-insensitively) and stored under the canonical roster name.
    If that cell is blank (e.g. a roster formula that didn't resolve), the row's
    position within the block falls back to the roster order.

    Returns ``(stages, union_columns, warnings)``.
    """
    roster_names = [r.get("Name", "") for r in roster]
    canonical = {n.lower(): n for n in roster_names if n}

    stages: list[Stage] = []
    union_columns: list[str] = []
    warnings: list[str] = []

    current: Stage | None = None       # active titled block (None between blocks)
    header: list[str] | None = None    # its column header row
    flag_cols: list[int] = []          # checkbox column indices for the current block
    skipping = False                   # inside an untitled (template placeholder) block
    row_index = 0                      # player position within the current block

    for raw in ws.iter_rows(values_only=True):
        cells = _cells(raw)
        raw_values = list(raw)         # keep original types (bool for checkboxes)

        if not any(cells):             # blank row ends the current block
            current, header, skipping = None, None, False
            continue

        first = cells[0]

        # Header row of a block.
        if first.lower() in _HEADER_LABELS:
            header = cells
            flag_cols = [i for i in range(1, len(cells)) if cells[i].lower() in _BOLD_LABELS]
            if current is None:
                skipping = True        # header with no title above -> dormant block
            else:
                current.columns = [c for i, c in enumerate(cells[1:], start=1) if c and i not in flag_cols]
                for col in current.columns:
                    if col not in union_columns:
                        union_columns.append(col)
                row_index = 0
            continue

        if skipping:
            continue

        # Player data row (active titled block with a header).
        if current is not None and header is not None:
            player = first
            if not player and row_index < len(roster_names):
                player = roster_names[row_index]      # positional fallback
            player = canonical.get(player.lower(), player)
            values = {}
            bold = False
            for col in range(1, len(header)):
                col_name = header[col]
                raw = raw_values[col] if col < len(raw_values) else None
                cell = cells[col] if col < len(cells) else ""
                # Any checkbox/boolean cell is the bold signal, never shown as gear.
                if col in flag_cols or isinstance(raw, bool) or cell.strip().lower() in ("true", "false"):
                    bold = bold or _truthy(raw if raw is not None else cell)
                    continue
                if col_name:
                    values[col_name] = cell
            values[_BOLD_KEY] = bold
            if player:
                current.rows[player] = values
            row_index += 1
            continue

        # Otherwise this is a stage-title row: start a new block.
        current = Stage(name=first, order=len(stages), columns=[], rows={})
        stages.append(current)
        header = None
        row_index = 0

    stages = [s for s in stages if s.rows]  # drop titled-but-empty blocks
    for i, stage in enumerate(stages):
        stage.order = i
    # Keep only columns that actually carry gear somewhere (drops stray/blank ones).
    used = {col for s in stages for vals in s.rows.values()
            for col in vals if col != _BOLD_KEY and vals.get(col)}
    union_columns = [c for c in union_columns if c in used]
    return stages, union_columns, warnings


def _cross_check(data: RaidData) -> list[str]:
    """Warn about names that appear in setups but not the roster, and vice versa."""
    warnings = []
    roster_lower = {row.get("Name", "").lower() for row in data.roster}
    setup_players = {p for stage in data.stages for p in stage.rows}
    for player in sorted(setup_players):
        if player.lower() not in roster_lower:
            warnings.append(f"'{player}' has setups but isn't in the Roster.")
    for name in data.player_names:
        if not any(_ci_get(stage.rows, name) for stage in data.stages):
            warnings.append(f"'{name}' is in the Roster but has no setups.")
    missing_tags = [row["Name"] for row in data.roster if not row.get("Discord", "").strip()]
    if missing_tags:
        warnings.append(
            "No Discord tag for: " + ", ".join(missing_tags)
            + " — they won't be able to run /setups until a tag is added."
        )
    return warnings


def load_raid(url_or_id: str) -> tuple[str, RaidData]:
    """Convenience: URL/ID → ``(sheet_id, RaidData)``. Raises ``SheetError`` on any problem."""
    sheet_id = extract_sheet_id(url_or_id)
    workbook = fetch_workbook(sheet_id)
    return sheet_id, parse_workbook(workbook)
